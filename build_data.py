"""
FSII viewer data builder — runs inside GitHub Actions.
 
What it does, unattended, on every run:
  1. Downloads the latest workbook from OneDrive (URL in the ONEDRIVE_URL secret),
     or reads a local file passed as argv[1] (for testing).
  2. Extracts the in-cell photos straight out of the .xlsx (richData chain) and
     resizes each to a web thumbnail.
  3. Reads the item data from the 'yacht 2 new product' tab.
  4. Encrypts everything (AES-256-GCM, PBKDF2-HMAC-SHA256, 600k iters) into data.enc.
 
Env vars (set as GitHub repo secrets):
  ONEDRIVE_URL  - a "view" (no-password) share link to the workbook  [required]
  GATE_USER     - viewer username  [required — never hardcoded here]
  GATE_PASS     - viewer password  [required — never hardcoded here]
 
Usage:
  python3 build_data.py                 # CI mode: download from ONEDRIVE_URL
  python3 build_data.py path/to.xlsx    # local test mode
"""
import sys, os, re, io, json, base64, zipfile, datetime, hashlib
import urllib.request, urllib.parse, http.cookiejar
 
from PIL import Image
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.ciphers.aead import AESGCM
 
ITER = 600000
# No real credentials live in this file — the repo is public. They are supplied
# only at runtime via the GATE_USER / GATE_PASS GitHub secrets. Missing = hard stop.
USER = os.environ.get('GATE_USER', '').strip().lower()
PASS = os.environ.get('GATE_PASS', '')
if not USER or not PASS:
    raise SystemExit('GATE_USER and GATE_PASS must be set in the environment '
                     '(GitHub secrets). Refusing to build without them.')
THUMB_MAX = 900
THUMB_Q = 82
 
ITEM_TYPE_MAP = {'SCULPTURE': 'Sculpture', 'SCULTPURE': 'Sculpture', 'STATUE': 'Sculpture',
                 'OBJECTS': 'Objects', 'SHELLS': 'Shells', 'DISHES': 'Dishes',
                 'LIGHTING': 'Lighting', 'PICTURES': 'Pictures', 'ARTWORKS': 'Artworks'}
 
 
def short_name(full):
    full = ' '.join(str(full).split())
    if len(full) <= 72:
        return full
    cut = full[:72]
    p = cut.rfind('. ')
    if p >= 30:
        return cut[:p + 1]
    w = cut.rfind(' ')
    base = cut[:w] if w > 40 else cut
    return base.rstrip(' ,;:') + '…'
 
 
def to_direct_download(url):
    """Turn a OneDrive / SharePoint share link into a direct-download URL."""
    if not url:
        return url
    # SharePoint / OneDrive-for-Business share links: append ?download=1
    if 'sharepoint.com' in url or '-my.sharepoint' in url or '1drv.ms' in url or 'onedrive.live.com' in url:
        if 'download=1' in url:
            return url
        sep = '&' if '?' in url else '?'
        return url + sep + 'download=1'
    return url
 
 
def _open_with_cookies(url):
    """SharePoint anonymous downloads bounce through a redirect that sets a
    guest-session cookie before serving the file. A plain urlopen drops that
    cookie and gets handed the HTML viewer page instead of the bytes. A
    cookie-aware opener that follows the whole redirect chain fixes it."""
    cj = http.cookiejar.CookieJar()
    opener = urllib.request.build_opener(
        urllib.request.HTTPCookieProcessor(cj),
        urllib.request.HTTPRedirectHandler())
    opener.addheaders = [
        ('User-Agent', 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) '
                       'AppleWebKit/537.36 (KHTML, like Gecko) '
                       'Chrome/124.0 Safari/537.36'),
        ('Accept', 'application/octet-stream,*/*')]
    with opener.open(url, timeout=180) as r:
        return r.read()
 
 
def fetch_workbook(dest):
    url = os.environ.get('ONEDRIVE_URL')
    if not url:
        raise SystemExit('ONEDRIVE_URL not set')
    dl = to_direct_download(url)
    data = None
    # Try the download=1 form first, then the raw link, with a cookie jar each.
    for candidate in (dl, url):
        try:
            data = _open_with_cookies(candidate)
        except Exception as e:
            data = None
            continue
        if data[:2] == b'PK':
            break  # got a real .xlsx (zip)
    if data is None:
        # Never surface the URL itself — public Actions logs.
        raise SystemExit('Download failed (network/redirect error). Check the '
                         'ONEDRIVE_URL secret is a valid no-password '
                         '"anyone with the link" link.')
    if data[:2] != b'PK':
        raise SystemExit('Download did not return an .xlsx (got a non-zip response, '
                         f'{len(data)} bytes) — SharePoint served a web page instead '
                         'of the file. The link opens anonymously but is not serving '
                         'the raw file to scripts.')
    open(dest, 'wb').write(data)
    return dest
 
 
def extract_incell_images(xlsx_path):
    """Return {row_number: image_bytes} by walking the richData chain."""
    z = zipfile.ZipFile(xlsx_path)
    wb = z.read('xl/workbook.xml').decode()
    rels = z.read('xl/_rels/workbook.xml.rels').decode()
    sid = {}
    for m in re.finditer(r'<sheet[^>]*name="([^"]+)"[^>]*r:id="(rId\d+)"', wb):
        sid[m.group(1)] = m.group(2)
    ridtarget = dict(re.findall(r'<Relationship Id="(rId\d+)"[^>]*Target="([^"]+)"', rels))
    tgt = ridtarget[sid['yacht 2 new product']]
    tgt = tgt[1:] if tgt.startswith('/') else 'xl/' + tgt
    sheet = z.read(tgt).decode()
 
    meta = z.read('xl/metadata.xml').decode()
    rvb = [int(m) for m in re.findall(r'<xlrd:rvb i="(\d+)"/>', meta)]
    rv = z.read('xl/richData/rdrichvalue.xml').decode()
    rvfirst = []
    for block in re.findall(r'<rv [^>]*>(.*?)</rv>', rv, re.S):
        vs = re.findall(r'<v>(.*?)</v>', block, re.S)
        rvfirst.append(int(vs[0]) if vs else None)
    rvrel = z.read('xl/richData/richValueRel.xml').decode()
    relids = re.findall(r'<rel r:id="(rId\d+)"/>', rvrel)
    rrels = z.read('xl/richData/_rels/richValueRel.xml.rels').decode()
    rid2media = dict(re.findall(r'<Relationship Id="(rId\d+)"[^>]*Target="([^"]+)"', rrels))
 
    row_vm = {int(r): int(vm) for r, vm in re.findall(r'<c r="[A-Z]+(\d+)"[^>]*vm="(\d+)"', sheet)}
    out = {}
    for row, vm in sorted(row_vm.items()):
        try:
            rvidx = rvb[vm - 1]
            relidx = rvfirst[rvidx]
            media = rid2media[relids[relidx]]
            out[row] = z.read('xl/' + media.replace('../', ''))
        except (IndexError, KeyError):
            continue
    return out
 
 
def make_thumb(raw):
    im = Image.open(io.BytesIO(raw))
    if im.mode not in ('RGB', 'L'):
        im = im.convert('RGB')
    im.thumbnail((THUMB_MAX, THUMB_MAX), Image.LANCZOS)
    buf = io.BytesIO()
    im.save(buf, 'JPEG', quality=THUMB_Q, optimize=True)
    return buf.getvalue()
 
 
def read_items(xlsx_path, row_imgs):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ss = wb['yacht 2 new product']
    rebuilt = ss.cell(row=1, column=1).value == 'Accession #'
    H2K = {'Description / Item Name': 'desc', 'Item Type': 'type', 'Dimensions': 'dims',
           'Weight (kg)': 'weight', 'Order Status': 'status', 'Deck #': 'deck',
           'Vitrine (Open or Closed)': 'vitrine', 'Current Location / Moved To': 'present',
           'Count': 'count'}
    if rebuilt:
        CM = {}
        for c in range(1, 30):
            h = ss.cell(row=1, column=c).value
            if h in H2K:
                CM[H2K[h]] = c
    else:
        CM = dict(desc=5, type=16, dims=6, weight=7, present=2, count=4)
 
    items = []
    for r in range(2, 200):
        d = ss.cell(row=r, column=CM['desc']).value
        if d in (None, ''):
            continue
        acc = f'FS-2026-{r - 1:03d}'
        traw = str(ss.cell(row=r, column=CM['type']).value or '').strip()
        typ = traw if rebuilt else ITEM_TYPE_MAP.get(traw.upper(), 'Objects')
        img = ''
        if r in row_imgs:
            img = 'data:image/jpeg;base64,' + base64.b64encode(make_thumb(row_imgs[r])).decode()
        _full = str(d).strip().title()
        _nm = short_name(_full)
        _base = _nm[:-1].rstrip() if _nm.endswith('…') else _nm
        _rest = ''
        if _nm != _full and _full.startswith(_base):
            _rest = _full[len(_base):].lstrip(' ,;:.').strip()
        items.append(dict(
            acc=acc, name=_nm, full=_full, rest=_rest, type=typ or 'Objects',
            dims=str(ss.cell(row=r, column=CM['dims']).value or ''),
            weight=ss.cell(row=r, column=CM['weight']).value if 'weight' in CM else None,
            deck=str(ss.cell(row=r, column=CM['deck']).value or '') if 'deck' in CM else '',
            vitrine=str(ss.cell(row=r, column=CM['vitrine']).value or '') if 'vitrine' in CM else '',
            status=str(ss.cell(row=r, column=CM['status']).value or '') if 'status' in CM else '',
            count=ss.cell(row=r, column=CM['count']).value if 'count' in CM else None,
            loc=str(ss.cell(row=r, column=CM['present']).value or '') if 'present' in CM else '',
            img=img))
    return items
 
 
def main():
    if len(sys.argv) > 1:
        xlsx = sys.argv[1]
    else:
        xlsx = 'workbook.xlsx'
        fetch_workbook(xlsx)
 
    # Change-detection: only rebuild data.enc when the workbook actually changed.
    # Otherwise the random salt/nonce would make data.enc differ every run and
    # bloat the repo with a new multi-MB commit every 30 minutes.
    src_hash = hashlib.sha256(open(xlsx, 'rb').read()).hexdigest()
    prev = ''
    if os.path.exists('.datahash'):
        prev = open('.datahash').read().strip()
    if src_hash == prev and os.path.exists('data.enc'):
        print('No change in source workbook — skipping rebuild.')
        return
 
    row_imgs = extract_incell_images(xlsx)
    items = read_items(xlsx, row_imgs)
    stamp = int(datetime.datetime.now(datetime.timezone.utc).timestamp() * 1000)
 
    payload = json.dumps({'epoch': stamp, 'items': items}).encode()
    salt, nonce = os.urandom(16), os.urandom(12)
    kdf = PBKDF2HMAC(algorithm=hashes.SHA256(), length=32, salt=salt, iterations=ITER)
    key = kdf.derive(f'{USER}:{PASS}'.encode())
    ct = AESGCM(key).encrypt(nonce, payload, None)
    open('data.enc', 'w').write(base64.b64encode(salt + nonce + ct).decode())
    open('.datahash', 'w').write(src_hash)
    n_img = sum(1 for i in items if i['img'])
    print(f'OK: {len(items)} items, {n_img} photos, data.enc '
          f'{round(os.path.getsize("data.enc") / 1e6, 2)} MB')
 
 
if __name__ == '__main__':
    main()
 
